import streamlit as st
import pandas as pd
import re
from datetime import datetime
import altair as alt
import io
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# PAGE CONFIG + HEADER
# ------------------------------------------------------------

st.set_page_config(
    page_title="🌼 WhatsApp Data Parser — Final Edition",
    layout="wide"
)

st.title("🌼 WhatsApp Data Parser — Final Edition")

st.markdown("""
<div style="
    padding: 15px;
    border-radius: 10px;
    background-color: #e8f5e9;
    border-left: 6px solid #66BB6A;
    font-size: 16px;
    margin-bottom: 20px;
">
<b>🔒 Your Data Is Safe</b><br>
This tool processes your WhatsApp messages <i>only on your device</i>.
Nothing is stored, uploaded, or shared.
</div>
""", unsafe_allow_html=True)

# ------------------------------------------------------------
# TIMESTAMP PARSING + MERGING
# ------------------------------------------------------------

TS_PATTERN = r"^\d{1,2}/\d{1,2}/\d{2,4},\s\d{1,2}:\d{2}"

TS_FORMATS = [
    "%d/%m/%y, %H:%M",
    "%d/%m/%Y, %H:%M",
    "%d/%m/%y, %I:%M %p",
    "%d/%m/%Y, %I:%M %p",
]

def parse_ts(ts_str):
    for fmt in TS_FORMATS:
        try:
            return datetime.strptime(ts_str, fmt)
        except:
            continue
    return None

def parse_whatsapp_line(line):
    if re.match(TS_PATTERN, line):
        try:
            ts_part, rest = line.split(" - ", 1)
            sender, msg = rest.split(": ", 1)
            return ts_part.strip(), sender.strip(), msg.strip()
        except:
            return None, None, None
    return None, None, None

def merge_messages(lines):
    merged = []
    current_ts = None
    current_sender = None
    current_msg = []

    for line in lines:
        ts_str, sender, msg = parse_whatsapp_line(line)
        if ts_str and sender and msg:
            if current_ts is not None:
                merged.append((current_ts, current_sender, "\n".join(current_msg)))
            current_ts = ts_str
            current_sender = sender
            current_msg = [msg]
        else:
            if current_ts is not None and line.strip():
                current_msg.append(line.strip())

    if current_ts is not None:
        merged.append((current_ts, current_sender, "\n".join(current_msg)))

    return merged

# ------------------------------------------------------------
# ADVANCED DYNAMIC TOKEN PARSING
# ------------------------------------------------------------

SYSTEM_IGNORE = {
    "THIS MESSAGE WAS DELETED",
    "MESSAGE WAS DELETED",
    "EDITED",
    "MEDIA OMITTED",
    "GROUP CLOSED",
    "GROUP OPEN",
    "REGISTRATION",
    "CHANGED THIS GROUP'S SETTINGS",
    "CANCEL"
}

def clean_token(token):
    cleaned = re.sub(r"[^\w]", "", token)
    return cleaned.strip()

def tokenize_dynamic(msg):
    raw_tokens = msg.split()
    tokens = []
    for t in raw_tokens:
        c = clean_token(t)
        if c:
            tokens.append(c)
    return tokens

def split_multi_person(msg):
    """
    Split ONLY on newline; ignore lines containing TOTAL.
    """
    parts = msg.split("\n")
    cleaned = []
    for p in parts:
        up = p.upper().strip()
        if "TOTAL" in up:
            continue
        if p.strip():
            cleaned.append(p.strip())
    return cleaned

def parse_people_from_message(msg):
    msg_upper = msg.upper()

    # Ignore system messages entirely
    if any(key in msg_upper for key in SYSTEM_IGNORE):
        return []

    persons = split_multi_person(msg)
    rows = []

    for person in persons:
        tokens = tokenize_dynamic(person)

        if not tokens:
            continue

        numeric = []
        alpha = []
        alnum = []

        for t in tokens:
            # Extract number even if mixed (e.g., 54PR)
            m = re.search(r"\d+", t)
            if m:
                numeric.append(m.group())  # number part

                leftover = re.sub(r"\d+", "", t)
                if leftover:
                    if leftover.isalpha():
                        alpha.append(leftover)
                    else:
                        alnum.append(leftover)
                continue

            # Pure alphabetic
            if t.isalpha():
                alpha.append(t)
                continue

            # Everything else
            alnum.append(t)

        ordered = numeric + alpha + alnum

        row = {}
        for i, tok in enumerate(ordered, start=1):
            row[f"Col{i}"] = tok

        rows.append(row)

    return rows

# ------------------------------------------------------------
# STREAMLIT UI
# ------------------------------------------------------------

def main():
    uploaded = st.file_uploader("Upload WhatsApp .txt file", type=["txt"])

    if uploaded:
        raw_lines = uploaded.read().decode("utf-8", errors="ignore").split("\n")
        merged = merge_messages(raw_lines)

        rows = []
        timestamps = []

        for ts_str, sender, msg in merged:
            dt = parse_ts(ts_str)
            if not dt:
                continue
            timestamps.append(dt)

            people = parse_people_from_message(msg)
            for p in people:
                base = {
                    "Sender": sender,
                    "Timestamp": dt,
                    "Raw Message": msg,
                }
                base.update(p)

                # Needs Review logic
                tokens = list(p.values())
                needs_review = False

                # 1. Col1 should be numeric
                if tokens and not tokens[0].isdigit():
                    needs_review = True

                # 2. Too few tokens
                if len(tokens) < 2:
                    needs_review = True

                # 3. Mixed leftover tokens (letters+numbers)
                for t in tokens:
                    if re.search(r"[A-Za-z]", t) and re.search(r"\d", t):
                        needs_review = True

                # 4. TOTAL in message
                if "TOTAL" in msg.upper():
                    needs_review = True

                base["Needs Review"] = "YES" if needs_review else "NO"

                rows.append(base)

        if rows:
            df = pd.DataFrame(rows)

            # ------------------------------------------------------------
            # DATE FILTERS
            # ------------------------------------------------------------
            st.subheader("Select Date & Time Range")
            col1, col2 = st.columns(2)

            start_date = col1.date_input("Start Date", min(timestamps))
            start_time = col1.time_input("Start Time", min(timestamps).time())

            end_date = col2.date_input("End Date", max(timestamps))
            end_time = col2.time_input("End Time", max(timestamps).time())

            start_dt = datetime.combine(start_date, start_time)
            end_dt = datetime.combine(end_date, end_time)

            mask = (df["Timestamp"] >= start_dt) & (df["Timestamp"] <= end_dt)
            df_filtered = df[mask].copy()

            st.subheader("Parsed Data")
            st.dataframe(df_filtered, use_container_width=True)

            # ------------------------------------------------------------
            # PRELIMINARY ANALYTICS (light)
            # ------------------------------------------------------------

            st.subheader("Preliminary Analysis Summary")

            cat_cols = []
            for col in df_filtered.columns:
                if col.startswith("Col"):
                    uniques = df_filtered[col].dropna().unique()
                    if 2 <= len(uniques) <= 5:
                        cat_cols.append(col)
            cat_cols = cat_cols[:2]

            if cat_cols:
                chart_cols = st.columns(2)

                for idx, col in enumerate(cat_cols):
                    vc = df_filtered[col].value_counts().reset_index()
                    vc.columns = [col, "Count"]

                    chart = alt.Chart(vc).mark_bar(
                        cornerRadiusTopLeft=6,
                        cornerRadiusTopRight=6
                    ).encode(
                        x=f"{col}:N",
                        y="Count:Q",
                        color=alt.value("#81C784")
                    )

                    chart_cols[idx].altair_chart(chart, use_container_width=True)

                st.markdown(
                    f"**Detected categorical fields:** {', '.join(cat_cols)} — useful for early pattern insights."
                )
            else:
                st.info("No suitable categorical columns found for preliminary charts.")

            # ------------------------------------------------------------
            # DOWNLOAD AS EXCEL
            # ------------------------------------------------------------

            def build_excel(df):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    # Parsed Data sheet
                    df.to_excel(writer, sheet_name="Parsed Data", index=False)

                    # Color formatting for Needs Review
                    if "Needs Review" in df.columns:
                        wb = writer.book
                        ws = writer.sheets["Parsed Data"]
                        col_idx = df.columns.get_loc("Needs Review") + 1
                        col_letter = get_column_letter(col_idx)
                        red_fill = PatternFill(
                            start_color="FFC7CE",
                            end_color="FFC7CE",
                            fill_type="solid"
                        )
                        for row in range(2, len(df) + 2):
                            cell = ws[f"{col_letter}{row}"]
                            if cell.value == "YES":
                                cell.fill = red_fill

                    # Reviewer Notes sheet
                    notes = [
                        ["Reviewer Notes"],
                        [""],
                        ["Some rows or columns may need human supervision:"],
                        ["1. Rows marked 'YES' in Needs Review column."],
                        ["2. Columns with mixed tokens (letters + numbers)."],
                        ["3. Rows where Col1 is not numeric (age or key missing)."],
                        ["4. Rows with fewer than 2 tokens (possible incomplete entry)."],
                        ["5. Rows containing unusual codes (PR, PS, NV, Veg, NonVeg, etc.)."],
                        ["6. Rows coming from multi-person messages (multiple lines in Raw Message)."],
                        ["7. Rows related to summary or TOTAL lines in the original chat."],
                    ]
                    notes_df = pd.DataFrame(notes)
                    notes_df.to_excel(
                        writer,
                        sheet_name="Reviewer Notes",
                        header=False,
                        index=False
                    )

                output.seek(0)
                return output

            excel_file = build_excel(df_filtered)

            st.download_button(
                label="📥 Download Excel File",
                data=excel_file,
                file_name="whatsapp_parsed.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# ------------------------------------------------------------
# ENTRYPOINT
# ------------------------------------------------------------

if __name__ == "__main__":
    main()
